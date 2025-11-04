import pdfplumber
import csv
import re
import argparse
import logging
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import PieChart, Reference

logging.basicConfig(level=logging.INFO)

# ... (TRADE_KEYWORDS, assign_trade, is_line_item, extract_xactimate_items as before) ...
TRADE_KEYWORDS = {
    "Floor Protection": ["floor protection", "cardboard", "protect floor", "mask floor"],
    "Insulation": ["insulation", "batt", "fiberglass", "blown-in"],
    "Drywall": ["drywall", "sheetrock", "tape joint", "texture", "patch", "mud", "repair wall", "joint compound", "corner bead"],
    "Painting": ["paint", "painting", "primer", "prime", "seal", "coating", "enamel", "mask wall", "caulk", "caulking", "mirror", "towel bar", "toilet paper", "tp holder"],
    "Baseboards, Trim, Casing": ["baseboard", "trim", "casing", "moulding", "crown", "shoe mould", "quarter round"],
    "Doors": ["door", "door stop", "interior door", "slab"],
    "Shower": ["shower", "shower pan", "shower door", "shower surround"],
    "Laminate": ["laminate", "pergo", "engineered wood"],
    "Vinyl Flooring": ["vinyl floor", "vinyl sheet"],
    "Tile Flooring": ["tile floor", "ceramic tile", "porcelain tile", "grout", "thinset", "cement board", "grout", "clean floor and prep"],
    "Carpet": ["carpet", "carpeting", "pad", "broadloom", "tack stip"],
    "HVAC": ["hvac", "register", "ventilation"],
    "Content Manipulation": ["content manipulation", "move contents", "protect contents", "cover contents", "contents"],
    "Cleaning": ["clean", "cleaning", "final clean", "clean up", "final cleanup", "final cleaning", "construction clean"],
    "Debris Removal": ["debris removal", "dump", "haul debris", "remove debris", "trash out"],
    "Cabinets": ["cabinet", "vanity", "base cabinet", "wall cabinet", "countertop"],
    "Electrical": ["electrical", "outlet", "switch", "receptacle", "breaker", "light fixture", "light", "ceiling fan"],
    "Showers, Tubs, Tile": ["tub", "bathtub", "shower", "tile", "surround", "enclosure", "shower pan", "shower door", "shower surround"],
    "Plumbing, Toilets, Sinks": ["plumbing", "toilet", "sink", "faucet", "supply line", "angle stop", "drain", "p-trap"],
    "Labor Minimums": ["labor minimum"],
    "Mitigation": ["water extraction", "remediation", "mitigation"]
}

HEADERS = ["DESCRIPTION", "TRADE", "QUANTITY", "UNIT PRICE", "TAX", "O&P", "RCV", "DEPREC.", "ACV"]

LINE_ITEM_REGEX = re.compile(
    r"^(\d+\.\s+.+?)\s+(\d+\.\d+\s+(?:SF|LF|EA|HR|DA|SY))\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)\s+\(([\d,.]+)\)\s+([\d,.]+)"
)


def assign_trade(description):
    desc = description.lower()
    for trade, keywords in TRADE_KEYWORDS.items():
        if any(keyword in desc for keyword in keywords):
            return trade
    return "Other"


def is_line_item(line):
    return bool(re.match(r"^\d+\.", line.strip()))


def should_skip_line(line):
    """
    Determine if a line should be skipped (not a line item).
    Returns True for dimensions, totals, notes, headers, etc.
    """
    line_lower = line.lower()

    # Skip dimension headers and dimension lines
    if 'dimension' in line_lower:
        return True

    # Skip total/subtotal lines
    if any(keyword in line_lower for keyword in ['total:', 'subtotal:', 'grand total', 'line item total']):
        return True

    # Skip section headers and notes
    if any(keyword in line_lower for keyword in ['**contents**', '**claim info**', '**summary**', 'estimate summary', 'adjuster summary']):
        return True

    # Skip lines that are just headers or instructions
    if any(keyword in line_lower for keyword in ['receipts must be', 'items on receipts', 'the receipt should contain', 'additional documentation']):
        return True

    # Skip lines with "SF" or other units but no leading number (these are dimension headers)
    if re.match(r'^\s*[A-Z\s]+(?:SF|LF|EA)\s*$', line, re.IGNORECASE):
        return True

    return False


def extract_xactimate_items(pdf_path):
    extracted_items = []

    # Multiple regex patterns to try in order (fallback logic)
    # Pattern 1: Full format with AGE/LIFE and CONDITION
    # Format: NUMBER. DESCRIPTION QTY+UNIT UNIT_PRICE TAX O&P RCV AGE/LIFE [yrs] Text COND% (DEPREC) ACV
    # Example: 1. Remove... 13.49SQ 7.27 0.00 9.80 107.87 9/NA Avg. 0% (0.00) 107.87
    pattern_with_age_life = re.compile(
        r"^(\d+\.)\s+"  # Line number with period
        r"(.+?)\s+"  # Description (non-greedy, allow special chars)
        r"([\d,.]+)([A-Z]{2,4})\s+"  # Quantity+unit combined (no space): 13.49SQ
        r"([\d,.]+)\s+"  # Unit price
        r"([\d,.]+)\s+"  # Tax
        r"([\d,.]+)\s+"  # O&P
        r"([\d,.]+)\s+"  # RCV
        r"[\d/NA]+\s+"  # AGE/LIFE (e.g., "9/30" or "9/NA" or "0/30")
        r"(?:yrs?\s+)?"  # Optional "yrs" or "yr"
        r"[A-Za-z.]+\s+"  # Text like "Avg."
        r"\d+(?:\.\d+)?%\s+"  # CONDITION percentage (e.g., "30%" or "25.71%")
        r"(?:\[M\]\s+)?"  # Optional depreciation marker [M]
        r"\(([\d,.]+)\)\s+"  # Depreciation in parentheses
        r"([\d,.]+)"  # ACV
        r"(?:\s|$)"  # End with whitespace or end of line
    )

    # Pattern 6: State Farm 3-line format (CONDITION on separate line AFTER deprec/ACV)
    # Format: NUMBER. DESCRIPTION QTY+UNIT UNIT_PRICE TAX O&P RCV AGE/LIFE [yrs] (DEPREC) ACV
    # Next line: Text COND%
    # Example: 2. Laminated... 19.67SQ 433.28 310.68 1,766.66 10,599.96 8/30 yrs (2,826.65) 7,773.31
    #          Avg. 26.67%
    pattern_state_farm = re.compile(
        r"^(\d+\.)\s+"  # Line number with period
        r"(.+?)\s+"  # Description (non-greedy, allow special chars)
        r"([\d,.]+)([A-Z]{2,4})\s+"  # Quantity+unit combined (no space): 19.67SQ
        r"([\d,.]+)\s+"  # Unit price
        r"([\d,.]+)\s+"  # Tax
        r"([\d,.]+)\s+"  # O&P
        r"([\d,.]+)\s+"  # RCV
        r"[\d/NA]+\s+"  # AGE/LIFE (e.g., "8/30")
        r"(?:yrs?\s+)?"  # Optional "yrs" or "yr"
        r"\(([\d,.]+)\)\s+"  # Depreciation in parentheses (NO CONDITION before this!)
        r"([\d,.]+)"  # ACV
        r"(?:\s|$)"  # End with whitespace or end of line
    )

    # Pattern 2: Simple format without AGE/LIFE columns
    # Format: NUMBER. DESCRIPTION QTY+UNIT UNIT_PRICE TAX O&P RCV (DEPREC) ACV
    # Example: 52. R&R Vinyl window... 3.00EA 895.87 195.90 288.36 3,171.87 (951.56) 2,220.31
    pattern_simple = re.compile(
        r"^(\d+\.)\s+"  # Line number with period
        r"(.+?)\s+"  # Description (non-greedy, allow special chars)
        r"([\d,.]+)([A-Z]{2,4})\s+"  # Quantity+unit combined (no space)
        r"([\d,.]+)\s+"  # Unit price
        r"([\d,.]+)\s+"  # Tax
        r"([\d,.]+)\s+"  # O&P
        r"([\d,.]+)\s+"  # RCV
        r"[\(<]([\d,.]+)[\)>]\s+"  # Depreciation (parentheses or angle brackets)
        r"([\d,.]+)"  # ACV
        r"(?:\s|$)"  # End with whitespace or end of line
    )

    # Pattern 3: Alternative format with angle brackets for depreciation
    # Format: NUMBER. DESCRIPTION QTY UNIT UNIT_PRICE TAX O&P RCV <DEPREC> ACV
    pattern_angle_brackets = re.compile(
        r"^(\d+\.)\s+"  # Line number with period
        r"(.+?)\s+"  # Description (non-greedy)
        r"([\d,.]+)\s+"  # Quantity
        r"([A-Z]{2,4})\s+"  # Unit (separate from quantity)
        r"([\d,.]+)\s+"  # Unit price
        r"([\d,.]+)\s+"  # Tax
        r"([\d,.]+)\s+"  # O&P
        r"([\d,.]+)\s+"  # RCV
        r"<([\d,.]+)>\s+"  # Depreciation in angle brackets
        r"([\d,.]+)"  # ACV
        r"(?:\s|$)"  # End with whitespace or end of line
    )

    # Pattern 4: No TAX/O&P columns (Allstate LS format)
    # Format: NUMBER. DESCRIPTION QTY+UNIT UNIT_PRICE RCV AGE/LIFE [yrs] Text COND% (DEPREC) ACV
    # Example: 19. Paint trim - one coat 18.00LF 1.06 19.08 0/15 yrs Avg. 0% (0.00) 19.08
    # Alt:     1. Remove Laminated... 13.74SQ 82.16 1,128.88 0/30 yrs Avg. NA (0.00) 1,128.88
    pattern_no_tax_op = re.compile(
        r"^(\d+\.)\s+"  # Line number with period
        r"(.+?)\s+"  # Description (non-greedy, allow special chars)
        r"([\d,.]+)([A-Z]{2,4})\s+"  # Quantity+unit combined (no space): 18.00LF
        r"([\d,.]+)\s+"  # Unit price
        r"([\d,.]+)\s+"  # RCV (no tax, no O&P!)
        r"[\d/NA]+\s+"  # AGE/LIFE (e.g., "0/15" or "6/30")
        r"(?:yrs?\s+)?"  # Optional "yrs" or "yr"
        r"[A-Za-z.]+\s+"  # Text like "Avg."
        r"(?:\d+(?:\.\d+)?%|NA)\s+"  # CONDITION: percentage (e.g., "0%" or "20%") OR "NA"
        r"(?:\[M\]\s+)?"  # Optional depreciation marker [M]
        r"[\(<]([\d,.]+)[\)>]\s+"  # Depreciation (parentheses or angle brackets)
        r"([\d,.]+)"  # ACV
        r"(?:\s|$)"  # End with whitespace or end of line
    )

    # Pattern 5: Has TAX but NO O&P (State Farm/Travelers multi-line)
    # Format: NUMBER. DESCRIPTION QTY+UNIT UNIT_PRICE TAX RCV AGE/LIFE [yrs] Text COND (DEPREC) ACV
    # Example: 1. Tandem axle dump... 1.00EA 325.65 0.00 325.65 10/NA Avg. NA (0.00) 325.65
    pattern_tax_no_op = re.compile(
        r"^(\d+\.)\s+"  # Line number with period
        r"(.+?)\s+"  # Description (non-greedy, allow special chars)
        r"([\d,.]+)([A-Z]{2,4})\s+"  # Quantity+unit combined (no space): 1.00EA
        r"([\d,.]+)\s+"  # Unit price
        r"([\d,.]+)\s+"  # TAX (has tax!)
        r"([\d,.]+)\s+"  # RCV (NO O&P!)
        r"[\d/NA]+\s+"  # AGE/LIFE (e.g., "10/NA" or "10/25")
        r"(?:yrs?\s+)?"  # Optional "yrs" or "yr"
        r"[A-Za-z.]+\s+"  # Text like "Avg."
        r"(?:\d+(?:\.\d+)?%|NA)\s+"  # CONDITION: percentage (e.g., "40%") OR "NA"
        r"(?:\[M\]\s+)?"  # Optional depreciation marker [M]
        r"[\(<]([\d,.]+)[\)>]\s+"  # Depreciation (parentheses or angle brackets)
        r"([\d,.]+)"  # ACV
        r"(?:\s|$)"  # End with whitespace or end of line
    )

    # List of patterns to try in order
    patterns = [
        ("with_age_life", pattern_with_age_life, "tax_op"),  # Has tax AND O&P, inline CONDITION
        ("state_farm", pattern_state_farm, "tax_op"),  # State Farm 3-line (CONDITION on line 3)
        ("tax_no_op", pattern_tax_no_op, "tax_only"),  # Has tax but NO O&P
        ("no_tax_op", pattern_no_tax_op, "no_tax_op"),  # NO tax, NO O&P
        ("simple", pattern_simple, "tax_op"),  # Has tax AND O&P
        ("angle_brackets", pattern_angle_brackets, "tax_op")  # Has tax AND O&P
    ]

    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, 1):
            text = page.extract_text()
            if not text:
                continue

            lines = text.split('\n')
            i = 0

            while i < len(lines):
                line = lines[i].strip()

                # Check if this line should be skipped
                if should_skip_line(line):
                    i += 1
                    continue

                # Detect start of a new item (number. ...)
                if re.match(r"^\d+\.\s", line):
                    # Check if this might be a multi-line format (State Farm/Travelers style)
                    # Where description is on one line and numbers are on the next
                    is_multiline = False
                    if i + 1 < len(lines):
                        next_line = lines[i + 1].strip()
                        # If next line starts with quantity+unit pattern (e.g., "1.00EA"), it's multi-line
                        if re.match(r"^[\d,.]+[A-Z]{2,4}\s", next_line):
                            is_multiline = True
                            combined_line = line + " " + next_line
                            j = i + 2  # Skip both lines
                        else:
                            # Standard multi-line combining
                            combined_line = line
                            j = i + 1

                            # Combine with following lines that are not new items
                            # Stop if we hit another numbered item, a skip pattern, or empty line
                            while j < len(lines):
                                next_line = lines[j].strip()

                                # Stop if new numbered item
                                if re.match(r"^\d+\.\s", next_line):
                                    break

                                # Stop if this is a skip pattern
                                if should_skip_line(next_line):
                                    break

                                # Stop if empty line
                                if not next_line:
                                    break

                                # Continue combining if it looks like a continuation
                                combined_line += " " + next_line
                                j += 1
                    else:
                        combined_line = line
                        j = i + 1

                    # Try each pattern in order until one matches
                    match = None
                    matched_pattern = None
                    has_tax_op = True

                    for pattern_name, pattern, has_tax in patterns:
                        match = pattern.match(combined_line)
                        if match:
                            matched_pattern = pattern_name
                            has_tax_op = has_tax
                            break

                    if match:
                        # Extract fields - pattern structure varies
                        line_num = match.group(1)
                        description = match.group(2).strip()
                        quantity = match.group(3)
                        unit = match.group(4)
                        unit_price = match.group(5)

                        # Special handling for State Farm 3-line format
                        # Check if next line (after combined_line) contains CONDITION percentage
                        if matched_pattern == "state_farm":
                            # j points to the line after the combined line
                            # Check if that line has "Avg. XX%"
                            if j < len(lines):
                                condition_line = lines[j].strip()
                                # Match lines like "Avg. 26.67%" or "Avg. 0%"
                                if re.match(r'^[A-Za-z.]+\s+\d+(?:\.\d+)?%', condition_line):
                                    # Found CONDITION line, skip it
                                    j += 1
                                # Also skip blank lines after CONDITION
                                while j < len(lines) and not lines[j].strip():
                                    j += 1

                        if has_tax_op == "tax_op":
                            # Patterns with TAX and O&P columns
                            tax = match.group(6)
                            o_p = match.group(7)
                            rcv = match.group(8)
                            deprec = match.group(9)
                            acv = match.group(10)
                        elif has_tax_op == "tax_only":
                            # Pattern with TAX but NO O&P (pattern_tax_no_op)
                            tax = match.group(6)
                            o_p = "0.00"  # No O&P column
                            rcv = match.group(7)
                            deprec = match.group(8)
                            acv = match.group(9)
                        else:  # "no_tax_op"
                            # Pattern without TAX/O&P (pattern_no_tax_op)
                            tax = "0.00"  # No tax column
                            o_p = "0.00"  # No O&P column
                            rcv = match.group(6)
                            deprec = match.group(7)
                            acv = match.group(8)

                        # Combine quantity and unit
                        quantity_unit = f"{quantity}{unit}"

                        # Clean up description (remove excessive whitespace)
                        description = re.sub(r'\s+', ' ', description)
                        description = f"{line_num} {description}"

                        # Assign trade category
                        trade = assign_trade(description)

                        extracted_items.append([
                            description, trade, quantity_unit, unit_price, tax, o_p, rcv, deprec, acv
                        ])
                    else:
                        # Only print NO MATCH for lines that start with numbers (potential line items)
                        if re.match(r"^\d+\.\s", combined_line):
                            print("NO MATCH (combined):", repr(combined_line[:200]))  # Truncate for readability

                    i = j  # Skip to next item
                else:
                    i += 1

    return [HEADERS] + extracted_items


def auto_fit_excel_columns(filename):
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
    wb.save(filename)


def bold_total_rows(filename):
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                # For trade sheets, look for "TOTAL" in the TRADE column
                if (cell.value == "TOTAL" or cell.value == "GRAND TOTAL"):
                    for c in row:
                        c.font = Font(bold=True)
                    break
    wb.save(filename)


def add_total_budget_rows(filename):
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        # Only apply to trade sheets (not Master or Totals)
        if ws.title in ["Master", "Totals"]:
            continue
        max_row = ws.max_row
        rcv_col = None
        # Find the RCV column (header row is 1)
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "RCV":
                rcv_col = idx
                break
        if not rcv_col:
            continue
        # Find the "TOTAL" row
        for row in range(2, max_row+1):
            if ws.cell(row=row, column=2).value == "TOTAL":
                total_row = row
                break
        else:
            continue
        # Add TOTAL BUDGET row
        budget_row = total_row + 1
        ws.cell(row=budget_row, column=2, value="TOTAL BUDGET")
        rcv_total_cell = f"{get_column_letter(rcv_col)}{total_row}"
        ws.cell(row=budget_row, column=rcv_col, value=f"={rcv_total_cell}*0.6")
        # Bold the TOTAL BUDGET row
        for c in ws[budget_row]:
            c.font = Font(bold=True)
    wb.save(filename)


def add_totals_pie_chart(filename):
    wb = load_workbook(filename)
    ws = wb["Totals"]
    max_row = ws.max_row

    # Find the RCV column index robustly
    rcv_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value and cell.value.strip().upper() == "RCV":
            rcv_col = idx
            break
    if not rcv_col:
        print("No RCV column found!")
        return

    # Exclude GRAND TOTAL row (last row)
    data = Reference(ws, min_col=rcv_col, min_row=1, max_row=max_row-1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=max_row-1)

    # Check data for debugging
    print("Pie chart data values:", [
          ws.cell(row=r, column=rcv_col).value for r in range(2, max_row)])
    print("Pie chart labels:", [ws.cell(row=r, column=1).value for r in range(2, max_row)])

    chart = PieChart()
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "RCV by Trade"
    chart.height = 10  # default is 7.5
    chart.width = 16  # default is 15

    # Place the chart offset rows below the table, in column C
    offset = 3
    chart_cell = f"C{max_row + offset}"
    ws.add_chart(chart, chart_cell)
    wb.save(filename)


def update_totals_with_budget(filename):
    wb = load_workbook(filename)
    ws = wb["Totals"]
    # Add "BUDGET" header
    ws.insert_cols(ws.max_column + 1)
    ws.cell(row=1, column=ws.max_column, value="BUDGET")
    for row in range(2, ws.max_row):  # skip header, skip GRAND TOTAL
        trade = ws.cell(row=row, column=1).value
        if trade and trade != "GRAND TOTAL":
            # Budget is always in ACV column, TOTAL BUDGET row (last row) of each trade sheet
            trade_ws = wb[trade[:31]]
            acv_col = None
            for idx, cell in enumerate(trade_ws[1], 1):
                if cell.value == "ACV":
                    acv_col = idx
                    break
            if not acv_col:
                continue
            # Find "TOTAL BUDGET" row
            for r in range(2, trade_ws.max_row+1):
                if trade_ws.cell(row=r, column=2).value == "TOTAL BUDGET":
                    budget_cell = f"'{trade[:31]}'!{get_column_letter(acv_col)}{r}"
                    ws.cell(row=row, column=ws.max_column, value=f"={budget_cell}")
                    break
    # Add grand total for BUDGET
    ws.cell(row=1, column=ws.max_column).value = "BUDGET"
    last_data_row = ws.max_row - 1
    ws.cell(row=ws.max_row, column=ws.max_column,
            value=f"=SUM({get_column_letter(ws.max_column)}2:{get_column_letter(ws.max_column)}{last_data_row})")
    wb.save(filename)


def save_to_excel_with_budget(data, excel_path):
    df = pd.DataFrame(data[1:], columns=data[0])

    # Convert numeric columns to float for calculations
    for col in ["UNIT PRICE", "TAX", "O&P", "RCV", "DEPREC.", "ACV"]:
        df[col] = df[col].replace(r'[\$,]', '', regex=True).astype(float)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Master sheet (all items)
        df.to_excel(writer, sheet_name="Master", index=False)

        # Trade sheets with TOTAL and TOTAL BUDGET rows
        trade_totals = []
        for trade, group in df.groupby('TRADE'):
            group = group.copy()
            # Calculate totals for numeric columns
            totals = {col: group[col].sum()
                      for col in ["UNIT PRICE", "TAX", "O&P", "RCV", "DEPREC.", "ACV"]}
            # TOTAL row
            total_row = [""] * len(group.columns)
            total_row[group.columns.get_loc("TRADE")] = "TOTAL"
            for col, val in totals.items():
                total_row[group.columns.get_loc(col)] = val
            # TOTAL BUDGET row (60% of RCV)
            budget_row = [""] * len(group.columns)
            budget_row[group.columns.get_loc("TRADE")] = "TOTAL BUDGET"
            # Add TOTAL and TOTAL BUDGET rows to DataFrame
            group = pd.concat([
                group,
                pd.DataFrame([total_row], columns=group.columns),
                pd.DataFrame([budget_row], columns=group.columns)
            ], ignore_index=True)
            group.to_excel(writer, sheet_name=trade[:31], index=False)
            # For Totals sheet
            trade_totals.append({
                "TRADE": trade,
                **{col: float(totals[col]) for col in ["UNIT PRICE", "TAX", "O&P", "RCV", "DEPREC.", "ACV"]}
            })

        # Totals sheet (summary of all trades)
        totals_df = pd.DataFrame(trade_totals)
        totals_df["RCV"] = pd.to_numeric(totals_df["RCV"], errors="coerce")
        # Add BUDGET column as 60% of RCV
        totals_df["BUDGET"] = totals_df["RCV"] * 0.6
        # Add GRAND TOTAL row
        grand_total = {"TRADE": "GRAND TOTAL"}
        for col in ["UNIT PRICE", "TAX", "O&P", "RCV", "DEPREC.", "ACV", "BUDGET"]:
            grand_total[col] = totals_df[col].sum()
        totals_df = pd.concat([totals_df, pd.DataFrame([grand_total])], ignore_index=True)
        totals_df.to_excel(writer, sheet_name="Totals", index=False)

    # Display key contractor information
    print("\n" + "="*60)
    print("CONTRACTOR SUMMARY")
    print("="*60)
    grand_total_row = totals_df[totals_df["TRADE"] == "GRAND TOTAL"].iloc[0]
    print(f"\n💰 INITIAL CHECK (ACV):        ${grand_total_row['ACV']:,.2f}")
    print(f"💵 TOTAL JOB VALUE (RCV):      ${grand_total_row['RCV']:,.2f}")
    print(f"⏳ HELD BACK (Depreciation):   ${grand_total_row['DEPREC.']:,.2f}")
    if grand_total_row['DEPREC.'] == 0:
        print("   ✅ No depreciation - Full replacement cost coverage")
    else:
        pct = (grand_total_row['DEPREC.'] / grand_total_row['RCV']) * 100
        print(f"   ⚠️  {pct:.1f}% of total held back as depreciation")
    print(f"\n📊 BUDGET (60% of RCV):        ${grand_total_row['BUDGET']:,.2f}")
    print("="*60)
    print("\nDetailed breakdown by trade:")
    print(totals_df[["TRADE", "RCV", "ACV", "DEPREC.", "BUDGET"]].to_string(index=False))

    # Auto-fit columns for all sheets
    auto_fit_excel_columns(excel_path)

    # Bold totals
    bold_total_rows(excel_path)

    # Add budget row
    add_total_budget_rows(excel_path)

    # Add hyperlinks
    # add_trade_hyperlinks(excel_path)

    # Add pie chart
    add_totals_pie_chart(excel_path)

    # Add budget totals to totals column
    # update_totals_with_budget(excel_path)

    logging.info(f"Saved Excel file with trades, master, and totals to {excel_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Extract Xactimate line items from PDF to Excel by trade with totals.")
    parser.add_argument("pdf_file", help="Path to the input PDF file")
    parser.add_argument("excel_file", help="Path to the output Excel file")
    args = parser.parse_args()

    logging.info(f"Extracting line items from {args.pdf_file} ...")
    data = extract_xactimate_items(args.pdf_file)
    if len(data) <= 1:
        logging.warning("No line items extracted.")
    else:
        save_to_excel_with_budget(data, args.excel_file)


if __name__ == "__main__":
    main()
